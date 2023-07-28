import os
from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import datetime


# 获取目录信息
def get_dirs_files(path, depth=0):
    data = []
    for entry in os.scandir(path):
        if entry.is_file():
            # 排除 *.html 的文件
            if str(entry.path).find(".html") == -1:
                data.append([depth, entry.path])
        elif entry.is_dir():
            data.append([depth, entry.path])
            data.extend(get_dirs_files(entry.path, depth + 1))
    return data


# 写入Excel
def write_to_excel(data, filename):
    wb = Workbook()
    ws = wb.active
    for row in data:
        depth = row[0]
        path = row[1]
        row_num = ws.max_row + 1
        for i in range(depth):
            cell = ws.cell(row=row_num, column=i + 1)
            cell.value = ''

        cell = ws.cell(row=row_num, column=depth + 1)
        cell.value = os.path.basename(path)
        cell.hyperlink = path
        red_font = Font(color="FFFF0000")
        if os.path.isdir(path):
            cell.font = red_font
    today = str(datetime.now()).replace(" ", "_").replace(":", "-").split(".")
    wb.save(f"{filename}_{today[0]}.xlsx")


data = get_dirs_files('E:/808/防盗信息/')
write_to_excel(data, '防盗信息')
